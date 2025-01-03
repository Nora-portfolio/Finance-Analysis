from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
CORS(app)  # This enables CORS for all routes

EXCEL_FILE = "data.xlsx"

@app.route('/')
def home():
    return "Welcome to the Flask App! Use the /save endpoint to send data."

# Initialize Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Amount", "Date", "Category"])  # Header row
    workbook.save(EXCEL_FILE)

@app.route('/save', methods=['POST'])
def save_data():
    data = request.json
    amount = data.get('amount')
    date = data.get('date')
    category = data.get('category')

    # Validate input
    if not amount or not date or not category:
        return jsonify({"message": "Invalid input! All fields are required."}), 400
    try:
        amount = float(amount)
        if amount <= 0:
            raise ValueError("Amount must be greater than zero.")
    except ValueError:
        return jsonify({"message": "Invalid amount! Must be a positive number."}), 400

    # Save to Excel
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([amount, date, category])
    workbook.save(EXCEL_FILE)

    return jsonify({"message": "Data saved successfully!"}), 200

if __name__ == '__main__':
    app.run(debug=True)


